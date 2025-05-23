from flask import Flask, request, jsonify
import sqlite3
import datetime
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font
from flask import send_file
from flask_cors import CORS
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
import shutil

app = Flask(__name__)
CORS(app)

DB_PATH = r'C:\Users\Gonzalo Bondar\Desktop\Ana Surak\sistema_ventas.db'
BACKUP_FOLDER = r'C:\Users\Gonzalo Bondar\Desktop\Ana Surak\backups'
LOG_BACKUP = os.path.join(BACKUP_FOLDER, "ultimo_backup.txt")  # Registro el último backup

def backup_db():
    """Realiza un backup de la base de datos solo una vez por día."""
    try:
        if not os.path.exists(BACKUP_FOLDER):
            os.makedirs(BACKUP_FOLDER)

        # Obtener la fecha actual
        hoy = datetime.datetime.now().strftime("%Y-%m-%d")

        # Revisar si ya se hizo un backup hoy
        if os.path.exists(LOG_BACKUP):
            with open(LOG_BACKUP, "r") as f:
                ultima_fecha = f.read().strip()
                if ultima_fecha == hoy:
                    print(f"⏳ Backup ya realizado hoy ({hoy}), no se repite.")
                    return  # Salir si ya hay backup del día

        # Nombre del backup basado en la fecha
        backup_path = os.path.join(BACKUP_FOLDER, f"backup_{hoy}.db")

        # Copiar la base de datos
        shutil.copy2(DB_PATH, backup_path)

        # Registrar que se hizo el backup
        with open(LOG_BACKUP, "w") as f:
            f.write(hoy)

        print(f"Backup realizado: {backup_path}")

    except Exception as e:
        print(f" Error en el backup: {e}")

# Llamar a la función de backup al iniciar el programa
backup_db()


#Conecta la BBDD por primera vez
def get_db_connection():
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row  
        print("Conexión exitosa a la base de datos")
        return conn
    except sqlite3.Error as e:
        print("Error al conectar a la base de datos:", e)
        return None

#Lee insumos
@app.route('/api/insumos', methods=['GET'])
def get_insumos():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id,nombre, cantidad, unidad_medida, precio_unitario FROM insumos")
        insumos = cursor.fetchall()
        conn.close()
        print("Datos obtenidos de la base de datos:", insumos)

        # Convierto a diccionario para JSON
        insumos_dict = [dict(row) for row in insumos]
        return jsonify(insumos_dict)
    except sqlite3.Error as e:
        print("Error en la consulta a la base de datos:", e)
        return jsonify({"error": "Error al obtener los datos de la base de datos"}), 500


#Agrega en tabla insumos
@app.route('/api/insumos', methods=['POST'])
def agregar_insumo():
    data = request.json
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO insumos (nombre, unidad_medida, cantidad, precio_unitario)
            VALUES (?, ?, ?, ?)
        """, (data['nombre'], data['unidad_medida'], data['cantidad'], data['precio_unitario']))
        conn.commit()
        conn.close()
        return jsonify({"message": "Insumo agregado correctamente"}), 201
    except sqlite3.Error as e:
        print("Error al agregar insumo:", e)
        return jsonify({"error": "Error al agregar insumo"}), 500

#Elimina de tabla insumos
@app.route('/api/insumos', methods=['DELETE'])
def eliminar_insumo():
    data = request.json
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM insumos WHERE nombre = ?", (data['nombre'],))
        conn.commit()
        conn.close()
        return jsonify({"message": "Insumo eliminado correctamente"}), 200
    except sqlite3.Error as e:
        print("Error al eliminar insumo:", e)
        return jsonify({"error": "Error al eliminar insumo"}), 500

#Modifica insumos
@app.route('/api/insumos', methods=['PUT'])
def modificar_insumo():
    data = request.json
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE insumos
            SET unidad_medida = ?, cantidad = ?, precio_unitario = ?
            WHERE nombre = ?
        """, (data['unidad_medida'], data['cantidad'], data['precio_unitario'], data['nombre']))
        conn.commit()
        conn.close()
        return jsonify({"message": "Insumo modificado correctamente"}), 200
    except sqlite3.Error as e:
        print("Error al modificar insumo:", e)
        return jsonify({"error": "Error al modificar insumo"}), 500
    
# Leer recetas
@app.route('/api/recetas', methods=['GET'])
def get_recetas():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, nombre, precio_venta, precio_venta_2 FROM recetas")
        recetas = cursor.fetchall()
        conn.close()
        return jsonify([dict(row) for row in recetas])
    except sqlite3.Error as e:
        return jsonify({"error": "Error al obtener recetas"}), 500


#fetch de insumo-recetas
    
@app.route('/api/receta_insumos/<int:receta_id>', methods=['GET'])
def get_receta_insumos(receta_id):
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.id AS insumo_id, i.nombre AS insumo, ri.cantidad, i.unidad_medida
            FROM receta_insumos ri
            JOIN insumos i ON ri.insumo_id = i.id
            WHERE ri.receta_id = ?
        """, (receta_id,))
        
        insumos = cursor.fetchall()
        conn.close()

        if not insumos:
            return jsonify([])  

        return jsonify([dict(row) for row in insumos])  

    except sqlite3.Error as e:
        return jsonify({"error": "Error al obtener los insumos de la receta"}), 500
    
#Delete de insumo-receta
@app.route('/api/receta_insumos/<int:receta_id>/<int:insumo_id>', methods=['DELETE'])
def eliminar_receta_insumo(receta_id, insumo_id):
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM receta_insumos WHERE receta_id = ? AND insumo_id = ?", (receta_id, insumo_id))
        conn.commit()
        conn.close()

        return jsonify({"message": "Insumo eliminado correctamente"}), 200

    except sqlite3.Error as e:
        return jsonify({"error": "Error al eliminar el insumo"}), 500


    
#Agregar insumo-receta
# Conversión de unidades (kg ↔ g, lt ↔ ml)
def convertir_unidad(cantidad, unidad_origen, unidad_destino):
    conversiones = {
        ("Kg", "Gr"): 1000,
        ("Gr", "Kg"): 0.001,
        ("Lt", "Ml"): 1000,
        ("Ml", "Lt"): 0.001,
        ("Unidades", "Unidades"): 1  
    }

    if (unidad_origen, unidad_destino) in conversiones:
        return cantidad * conversiones[(unidad_origen, unidad_destino)]
    else:
        return None  

@app.route('/api/receta_insumos/<int:receta_id>', methods=['PUT'])
def actualizar_receta_insumos(receta_id):
    data = request.json  

    print("🔹 Receta ID recibido:", receta_id)
    print("🔹 Datos recibidos:", data)

    if not isinstance(data, list):  
        return jsonify({"error": "Los datos deben ser una lista de insumos"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        for insumo in data:
            insumo_id = insumo.get("insumo_id")
            cantidad = insumo.get("cantidad")
            unidad_medida = insumo.get("unidad_medida")

            if not all([insumo_id, cantidad, unidad_medida]):
                continue  

            print(f"Procesando insumo: ID={insumo_id}, Cantidad={cantidad}, Unidad={unidad_medida}")

            # Obtener la unidad base desde la base de datos
            cursor.execute("SELECT unidad_medida FROM insumos WHERE id = ?", (insumo_id,))
            resultado = cursor.fetchone()
            if not resultado:
                print(f"❌ Insumo {insumo_id} no encontrado en la base de datos.")
                continue  # Insumo no encontrado

            unidad_base = resultado["unidad_medida"]

            # Convertir cantidad si la unidad ingresada es diferente a la base
            if unidad_medida != unidad_base:
                cantidad_convertida = convertir_unidad(cantidad, unidad_medida, unidad_base)
                if cantidad_convertida is None:
                    print(f"❌ No se pudo convertir {unidad_medida} a {unidad_base}")
                    return jsonify({"error": f"No se puede convertir {unidad_medida} a {unidad_base}"}), 400
            else:
                cantidad_convertida = cantidad

            # Insertar o actualizar el insumo en la receta
            cursor.execute("""
                INSERT INTO receta_insumos (receta_id, insumo_id, cantidad)
                VALUES (?, ?, ?)
                ON CONFLICT(receta_id, insumo_id) DO UPDATE SET cantidad = ?
            """, (receta_id, insumo_id, cantidad_convertida, cantidad_convertida))

        conn.commit()
        conn.close()
        print("✅ Datos insertados correctamente en receta_insumos.")
        return jsonify({"message": "Receta actualizada correctamente"}), 200

    except sqlite3.Error as e:
        conn.rollback()
        print("❌ Error en la base de datos:", e)
        return jsonify({"error": "Error al actualizar insumos"}), 500

#LEER DATOS

@app.route('/api/insumos/<nombre>', methods=['GET'])
def get_insumo(nombre):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Normalizar el nombre eliminando espacios extra y convirtiéndolo a minúsculas
    cursor.execute("SELECT nombre, cantidad, unidad_medida, precio_unitario FROM insumos WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?))", (nombre,))
    
    insumo = cursor.fetchone()
    conn.close()

    if insumo:
        return jsonify(dict(insumo))
    else:
        return jsonify({"error": "Insumo no encontrado"}), 404
    
# Agregar receta
@app.route('/api/recetas', methods=['POST'])
def agregar_receta():
    data = request.json
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO recetas (nombre, precio_venta, precio_venta_2)
            VALUES (?, ?, ?)
        """, (data['nombre'], data['precio_venta'], data.get('precio_venta_2', 0)))  # Usa 0 si no se envía
        conn.commit()
        conn.close()
        return jsonify({"message": "Receta agregada correctamente"}), 201
    except sqlite3.Error as e:
        return jsonify({"error": "Error al agregar receta"}), 500

# Modificar receta (nombre o precios)
@app.route('/api/recetas/<int:receta_id>', methods=['PUT'])
def modificar_receta(receta_id):
    data = request.json
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE recetas
            SET nombre = ?, precio_venta = ?, precio_venta_2 = ?
            WHERE id = ?
        """, (data['nombre'], data['precio_venta'], data['precio_venta_2'], receta_id))
        conn.commit()
        conn.close()
        return jsonify({"message": "Receta modificada correctamente"}), 200
    except sqlite3.Error as e:
        return jsonify({"error": "Error al modificar receta"}), 500


# Eliminar receta
@app.route('/api/recetas/<int:receta_id>', methods=['DELETE'])
def eliminar_receta(receta_id):
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM recetas WHERE id = ?", (receta_id,))
        conn.commit()
        conn.close()
        return jsonify({"message": "Receta eliminada correctamente"}), 200
    except sqlite3.Error as e:
        return jsonify({"error": "Error al eliminar receta"}), 500

#AGREGA STOCK Y PONDERA PRECIOS INGREDIENTES

@app.route('/api/insumos/actualizar_stock', methods=['PUT'])
def modificar_insumo_stock():
    data = request.json
    nombre = data['nombre']
    cantidad_nueva = data['cantidad']
    precio_nuevo = data['precio_unitario']

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Obtener el estado actual del insumo
        cursor.execute("SELECT cantidad, precio_unitario FROM insumos WHERE nombre = ?", (nombre,))
        insumo = cursor.fetchone()
        
        if insumo is None:
            return jsonify({"error": "Insumo no encontrado"}), 404
        
        cantidad_anterior, precio_anterior = insumo

        # Calcular el nuevo stock y el precio promedio ponderado
        nueva_cantidad = cantidad_anterior + cantidad_nueva
        nuevo_precio = ((cantidad_anterior * precio_anterior) + (cantidad_nueva * precio_nuevo)) / nueva_cantidad

        # Actualizar la base de datos
        cursor.execute("""
            UPDATE insumos
            SET cantidad = ?, precio_unitario = ?
            WHERE nombre = ?
        """, (nueva_cantidad, nuevo_precio, nombre))

        conn.commit()
        conn.close()
        
        return jsonify({
            "message": "Stock actualizado correctamente",
            "nombre": nombre,
            "nueva_cantidad": nueva_cantidad,
            "nuevo_precio": nuevo_precio
        }), 200

    except sqlite3.Error as e:
        print("Error al modificar stock:", e)
        return jsonify({"error": "Error al modificar stock"}), 500
    
# Agregar insumo a una receta
@app.route('/api/recetas/agregar_insumo', methods=['POST'])
def agregar_insumo_a_receta():
    data = request.json
    receta_id = data.get('receta_id')
    insumo_id = data.get('insumo_id')
    cantidad_ingresada = data.get('cantidad')
    unidad_ingresada = data.get('unidad_medida')

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()

        # Obtener la unidad de medida original del insumo
        cursor.execute("SELECT unidad_medida FROM insumos WHERE id = ?", (insumo_id,))
        insumo = cursor.fetchone()

        if not insumo:
            return jsonify({"error": "Insumo no encontrado"}), 404

        unidad_original = insumo[0]

        # Conversión de unidades
        conversiones = {
            ("g", "kg"): 1000,   # 1000 g = 1 kg
            ("kg", "g"): 1 / 1000,
            ("ml", "l"): 1000,   # 1000 ml = 1 l
            ("l", "ml"): 1 / 1000
        }

        if unidad_ingresada != unidad_original:
            clave_conversion = (unidad_ingresada.lower(), unidad_original.lower())
            if clave_conversion in conversiones:
                cantidad_convertida = cantidad_ingresada / conversiones[clave_conversion]
            else:
                return jsonify({"error": "Conversión no soportada"}), 400
        else:
            cantidad_convertida = cantidad_ingresada

        # Insertar en la tabla receta_insumos con la cantidad convertida
        cursor.execute("""
            INSERT INTO receta_insumos (receta_id, insumo_id, cantidad)
            VALUES (?, ?, ?)
        """, (receta_id, insumo_id, cantidad_convertida))

        conn.commit()
        conn.close()

        return jsonify({
            "message": "Insumo agregado a la receta correctamente",
            "receta_id": receta_id,
            "insumo_id": insumo_id,
            "cantidad_guardada": cantidad_convertida,
            "unidad_original": unidad_original
        }), 201

    except sqlite3.Error as e:
        print("Error al agregar insumo a receta:", e)
        return jsonify({"error": "Error en la base de datos"}), 500
    
# Fetch para ventas   
@app.route('/api/ventas', methods=['GET'])
def get_ventas():
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    if not desde or not hasta:
        return jsonify({"error": "Debe proporcionar un rango de fechas válido"}), 400

    # 🔹 Formatear fechas correctamente para SQLite
    desde = desde.replace("T", " ")
    hasta = hasta.replace("T", " ")

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()

        # 🔹 Nueva consulta incluyendo `v.id` como `venta_id`
        cursor.execute("""
            SELECT v.id AS venta_id,  -- ✅ Se incluye el ID de la venta
                   v.fecha_venta, 
                   SUM(dv.unidades) AS unidades_totales, 
                   SUM(dv.subtotal) AS total
            FROM ventas v
            JOIN detalle_ventas dv ON v.id = dv.venta_id
            WHERE datetime(v.fecha_venta) BETWEEN datetime(?) AND datetime(?)
            GROUP BY v.id
            ORDER BY v.fecha_venta ASC
        """, (desde, hasta))

        ventas = cursor.fetchall()
        conn.close()

        # 🔹 Transformar los resultados en una lista de diccionarios con formato de fecha
        ventas_list = []
        for row in ventas:
            ventas_list.append({
                "venta_id": row["venta_id"],  
                "fecha": row["fecha_venta"],
                "unidades_totales": row["unidades_totales"],
                "total": row["total"]
            })

        return jsonify(ventas_list)
    except sqlite3.Error as e:
        return jsonify({"error": "Error al obtener las ventas"}), 500

    
#Get para ventas 2
    
@app.route('/api/detalle_ventas', methods=['GET'])
def get_detalle_ventas():
    venta_id = request.args.get('venta_id')

    if not venta_id:
        return jsonify({"error": "Se requiere un venta_id"}), 400

    try:
        conn = get_db_connection()
        detalles = conn.execute("""
            SELECT dv.receta_id, r.nombre AS receta_nombre, dv.unidades, dv.precio_venta, dv.subtotal
            FROM detalle_ventas dv
            JOIN recetas r ON dv.receta_id = r.id
            WHERE dv.venta_id = ?;
        """, (venta_id,)).fetchall()
        conn.close()

        return jsonify([dict(row) for row in detalles])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

#Agregar nueva venta a tabla ventas
@app.route('/api/ventas', methods=['POST'])
def agregar_venta():
    data = request.json  # Recibe los datos en formato JSON

    fecha_venta = data.get("fecha_venta")
    medio_venta = data.get("medio_venta")
    total_venta = data.get("total_venta")

    if not fecha_venta or not medio_venta or total_venta is None:
        return jsonify({"error": "Datos incompletos"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Insertar la nueva venta en la tabla `ventas`
        cursor.execute("""
            INSERT INTO ventas (fecha_venta, medio_venta, total)
            VALUES (?, ?, ?)
        """, (fecha_venta, medio_venta, total_venta))

        conn.commit()
        venta_id = cursor.lastrowid  # Obtener el ID de la venta insertada
        conn.close()

        return jsonify({"message": "Venta agregada exitosamente", "venta_id": venta_id}), 201

    except sqlite3.Error as e:
        return jsonify({"error": "Error al insertar la venta"}), 500
    
#Post method para insertar en detalle_ventas
@app.route('/api/detalle_ventas', methods=['POST'])
def agregar_detalle_ventas():
    data = request.json  # Recibe los datos en formato JSON

    venta_id = data.get("venta_id")
    detalles = data.get("detalles")  # Lista de detalles de la venta

    if not venta_id or not detalles or not isinstance(detalles, list):
        return jsonify({"error": "Datos incompletos o formato incorrecto"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Insertar cada detalle en la tabla `detalle_ventas`
        for detalle in detalles:
            receta_id = detalle.get("receta_id")
            unidades = detalle.get("unidades")
            precio_venta = detalle.get("precio_venta")

            if not receta_id or not unidades or not precio_venta:
                return jsonify({"error": "Datos faltantes en los detalles"}), 400

            cursor.execute("""
                INSERT INTO detalle_ventas (venta_id, receta_id, unidades, precio_venta)
                VALUES (?, ?, ?, ?)
            """, (venta_id, receta_id, unidades, precio_venta))

        conn.commit()
        conn.close()

        return jsonify({"message": "Detalles de venta agregados exitosamente"}), 201

    except sqlite3.Error as e:
        return jsonify({"error": "Error al insertar detalles de venta"}), 500
    
# PUT para actualizar o agregar cliente
@app.route("/api/clientes", methods=["PUT"])
def upsert_cliente():
    data = request.json
    nombre = data.get("nombre", "").strip()
    celular = data.get("celular", "").strip()

    if not nombre and not celular:
        return jsonify({"message": "No se ingresó cliente, no se registrará"}), 200  

    if not nombre:
        nombre = "ANON"  # Si no hay nombre, lo registramos como ANON

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verificar si el cliente ya existe por número de celular
    cursor.execute("SELECT id FROM clientes WHERE celular = ?", (celular,))
    cliente = cursor.fetchone()

    if cliente:
        cliente_id = cliente["id"]
    else:
        #  Insertar nuevo cliente
        cursor.execute("INSERT INTO clientes (nombre, celular) VALUES (?, ?)", (nombre, celular))
        cliente_id = cursor.lastrowid  # Obtener el ID recién generado

    conn.commit()
    conn.close()

    return jsonify({"cliente_id": cliente_id}), 200  # Devuelve el ID del cliente registrado


#  PUT para asociar venta con cliente en cliente_ventas
@app.route("/api/cliente_ventas", methods=["PUT"])
def asociar_cliente_venta():
    data = request.json
    cliente_id = data.get("cliente_id")
    venta_id = data.get("venta_id")

    if not cliente_id or not venta_id:
        return jsonify({"error": "Faltan datos para la asociación"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    # Insertar relación en cliente_ventas si no existe
    cursor.execute(
        "INSERT OR IGNORE INTO cliente_ventas (cliente_id, venta_id) VALUES (?, ?)",
        (cliente_id, venta_id)
    )

    conn.commit()
    conn.close()

    return jsonify({"message": "Venta asociada al cliente correctamente"}), 200

@app.route('/api/insumos/update_bulk', methods=['PUT'])
def actualizar_insumos():
    data = request.json  # Recibe una lista de objetos {insumo_id, cantidad}
    
    if not data or not isinstance(data, list):
        return jsonify({"error": "Datos inválidos para actualizar insumos"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500
    
    try:
        cursor = conn.cursor()

        for item in data:
            # Validar que los datos sean correctos
            if 'insumo_id' not in item or 'cantidad' not in item:
                print("❌ Error: Falta 'insumo_id' o 'cantidad' en:", item)
                return jsonify({"error": "Faltan datos en la actualización de insumos"}), 400

            # Actualizar el stock del insumo sin permitir valores negativos
            cursor.execute("""
                UPDATE insumos
                SET cantidad = CASE 
                    WHEN cantidad - ? < 0 THEN 0 
                    ELSE cantidad - ? 
                END
                WHERE id = ?
            """, (item['cantidad'], item['cantidad'], item['insumo_id']))

        conn.commit()
        conn.close()
        return jsonify({"message": "✅ Insumos actualizados correctamente"}), 200

    except sqlite3.Error as e:
        print("❌ Error al actualizar insumos:", e)
        return jsonify({"error": "Error al actualizar insumos"}), 500

    
# DELETE para eliminar una venta y sus detalles
@app.route("/api/ventas/<int:venta_id>", methods=["DELETE"])
def eliminar_venta(venta_id):
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "No se pudo conectar a la base de datos"}), 500

    try:
        cursor = conn.cursor()

        # 1️⃣ Eliminar detalles de la venta
        cursor.execute("DELETE FROM detalle_ventas WHERE venta_id = ?", (venta_id,))

        # 2️⃣ Eliminar la venta principal
        cursor.execute("DELETE FROM ventas WHERE id = ?", (venta_id,))

        conn.commit()
        conn.close()

        return jsonify({"message": f"Venta {venta_id} eliminada correctamente"}), 200

    except sqlite3.Error as e:
        print("❌ Error al eliminar la venta:", e)
        return jsonify({"error": "Error al eliminar la venta"}), 500


#REPORTES 

@app.route('/api/reporte_ventas', methods=['GET'])
def generar_reporte_ventas():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener parámetros de fecha del request
    fecha_desde = request.args.get('desde')
    fecha_hasta = request.args.get('hasta')

    if not fecha_desde or not fecha_hasta:
        return jsonify({"error": "Se requieren ambas fechas para generar el reporte"}), 400

    try:
        # 🔹 Obtener detalles de ventas dentro del rango de fechas
        cursor.execute("""
            SELECT dv.venta_id, v.fecha_venta, r.nombre AS receta_nombre, 
                   dv.unidades, dv.precio_venta, v.medio_venta
            FROM detalle_ventas dv
            JOIN recetas r ON dv.receta_id = r.id
            JOIN ventas v ON dv.venta_id = v.id
            WHERE DATE(v.fecha_venta) BETWEEN DATE(?) AND DATE(?)
        """, (fecha_desde, fecha_hasta))
        detalles = cursor.fetchall()

        if not detalles:
            return jsonify({"error": "No hay datos disponibles en el rango seleccionado"}), 404

        # 🔹 Crear archivo Excel
        wb = openpyxl.Workbook()

        # ======== 📌 HOJA 1: DETALLE VENTAS ===========
        ws = wb.active
        ws.title = "Detalle Ventas"

        
        headers = ["Venta", "Fecha", "ID", "Unidades", "Medio de Venta", 
                   "Costo Total Insumos", "Valor Venta", "Comisiones Venta",
                   "Ganancia Bruta", "Rentabilidad Bruta"]
        ws.append(headers)

        for cell in ws[1]:  
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 🔹 Llenar datos
        medio_venta_count = {}  # Para la hoja de gráfico
        total_costos = total_venta = total_comisiones = total_ganancia = 0

        for row in detalles:
            venta_id, fecha_venta, receta_nombre, unidades, precio_venta, medio_venta = row

            # Convertir fecha al formato DD-MM-YYYY
            fecha_venta = fecha_venta.split("T")[0]
            fecha_venta = "-".join(reversed(fecha_venta.split("-")))

            # Obtener costos de insumos
            cursor.execute("""
                SELECT ri.insumo_id, ri.cantidad, i.precio_unitario
                FROM receta_insumos ri
                JOIN insumos i ON ri.insumo_id = i.id
                WHERE ri.receta_id = (SELECT id FROM recetas WHERE nombre = ?)
            """, (receta_nombre,))
            insumos = cursor.fetchall()

            costo_total_insumos = sum(insumo[1] * insumo[2] for insumo in insumos)
            comision_venta = 0.3 if medio_venta.lower() == "pedidosya" else 0
            ganancia_bruta = (precio_venta - costo_total_insumos) - (precio_venta * comision_venta)
            rentabilidad_bruta = (ganancia_bruta / costo_total_insumos) if costo_total_insumos > 0 else 0

            # Acumuladores para la fila total
            total_costos += costo_total_insumos * unidades
            total_venta += precio_venta * unidades
            total_comisiones += (comision_venta * precio_venta) * unidades
            total_ganancia += ganancia_bruta * unidades

            # Contar medios de venta para la hoja de gráficos
            if medio_venta in medio_venta_count:
                medio_venta_count[medio_venta] += unidades
            else:
                medio_venta_count[medio_venta] = unidades

            # Añadir cada unidad como fila separada
            for _ in range(unidades):
                row_data = [
                    receta_nombre, fecha_venta, venta_id, 1, medio_venta,
                    round(costo_total_insumos, 2), round(precio_venta, 2),
                    round(comision_venta * precio_venta, 2), round(ganancia_bruta, 2),
                    rentabilidad_bruta
                ]
                ws.append(row_data)

        # 🔹 Aplicar formato de porcentaje a la columna "Rentabilidad Bruta"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=10):
            for cell in row:
                cell.number_format = "0.00%"

        # 🔹 Fila Total
        last_row = ws.max_row + 1
        ws.append(["", "", "", "TOTAL:", "", 
                   round(total_costos, 2), round(total_venta, 2),
                   round(total_comisiones, 2), round(total_ganancia, 2),
                   f"=I{last_row}/F{last_row}"])

        for cell in ws[last_row]:
            cell.font = Font(bold=True, size=12)

        ws[f"J{last_row}"].number_format = "0.00%"  # Formato de porcentaje en total rentabilidad

        # 🔹 Ajustar tamaño de columnas automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # 🔹 Agregar filtro a la tabla
        ws.auto_filter.ref = f"A1:J{ws.max_row}"

        # ======== 📌 HOJA 2: MEDIO DE VENTA ===========
        ws_chart = wb.create_sheet(title="Medio de Venta")

        ws_chart.append(["Medio de Venta", "Cantidad de Ventas"])
        for medio, count in medio_venta_count.items():
            ws_chart.append([medio, count])

        # 🔹 Crear gráfico de torta
        pie = PieChart()
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=len(medio_venta_count) + 1)
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(medio_venta_count) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribución de Medios de Venta"
        pie.width = 18
        pie.height = 12

        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True

        pie.legend.position = "r"
        ws_chart.add_chart(pie, "D6")

        # 🔹 Guardar el archivo Excel
        file_path = f"reporte_ventas_{fecha_desde}_a_{fecha_hasta}.xlsx"
        wb.save(file_path)

        conn.close()

        return send_file(file_path, as_attachment=True)

    except sqlite3.Error as e:
        print("❌ Error al generar el reporte:", e)
        return jsonify({"error": "Error al generar el reporte"}), 500

    
@app.route('/api/reporte_clientes', methods=['GET'])
def generar_reporte_clientes():
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    if not desde or not hasta:
        return jsonify({"error": "Debe proporcionar un rango de fechas válido"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 🔹 Obtener clientes con sus datos
        cursor.execute("""
            SELECT c.id AS cliente_id, c.nombre, c.celular
            FROM clientes c
            JOIN cliente_ventas cv ON c.id = cv.cliente_id
            JOIN ventas v ON cv.venta_id = v.id
            WHERE date(v.fecha_venta) BETWEEN date(?) AND date(?)
            GROUP BY c.id
        """, (desde, hasta))
        clientes = cursor.fetchall()

        if not clientes:
            return jsonify({"error": "No hay clientes con ventas en el rango de fechas seleccionado"}), 404

        # 🔹 Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Clientes"

        # 🔹 Encabezados
        headers = ["Cliente", "Teléfono", "Cant. Ventas", "Monto Ventas"]
        ws.append(headers)

        for cell in ws[1]:  # Estilo para los encabezados
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 🔹 Llenar datos de clientes
        for cliente in clientes:
            cliente_id, nombre, celular = cliente

            # Obtener cantidad de ventas
            cursor.execute("""
                SELECT COUNT(DISTINCT venta_id)
                FROM cliente_ventas
                WHERE cliente_id = ?
            """, (cliente_id,))
            cant_ventas = cursor.fetchone()[0] or 0

            # Obtener monto total de ventas
            cursor.execute("""
                SELECT SUM(dv.subtotal)
                FROM cliente_ventas cv
                JOIN detalle_ventas dv ON cv.venta_id = dv.venta_id
                WHERE cv.cliente_id = ?
            """, (cliente_id,))
            monto_ventas = cursor.fetchone()[0] or 0

            ws.append([nombre, celular, cant_ventas, round(monto_ventas, 2)])

        # 🔹 Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # 🔹 Guardar el archivo con fecha en el nombre
        file_path = f"reporte_clientes_{desde}_a_{hasta}.xlsx"
        wb.save(file_path)

        conn.close()
        return send_file(file_path, as_attachment=True)

    except sqlite3.Error as e:
        print("❌ Error al generar el reporte de clientes:", e)
        return jsonify({"error": "Error al generar el reporte"}), 500
    
@app.route('/api/reporte_pedidos', methods=['GET'])
def generar_reporte_pedidos():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener parámetros de fecha del request
    fecha_desde = request.args.get('desde')
    fecha_hasta = request.args.get('hasta')

    if not fecha_desde or not fecha_hasta:
        return jsonify({"error": "Se requieren ambas fechas para generar el reporte"}), 400

    try:
        # 🔹 Obtener pedidos agrupados por receta_id dentro del rango de fechas
        cursor.execute("""
            SELECT r.nombre AS receta_nombre, 
                   SUM(dv.unidades) AS total_pedidos, 
                   SUM(dv.precio_venta * dv.unidades) AS total_ingreso
            FROM detalle_ventas dv
            JOIN recetas r ON dv.receta_id = r.id
            JOIN ventas v ON dv.venta_id = v.id
            WHERE DATE(v.fecha_venta) BETWEEN DATE(?) AND DATE(?)
            GROUP BY dv.receta_id
        """, (fecha_desde, fecha_hasta))
        
        pedidos = cursor.fetchall()

        if not pedidos:
            return jsonify({"error": "No hay datos disponibles en el rango seleccionado"}), 404

        # 🔹 Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Pedidos"

        # 🔹 Escribir encabezados con formato
        headers = ["Receta", "Total Pedidos", "Total Ingreso"]
        ws.append(headers)

        for cell in ws[1]:  
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        total_pedidos = 0
        total_ingreso = 0

        # 🔹 Llenar datos en la tabla
        for row in pedidos:
            receta_nombre, total_pedidos_receta, total_ingreso_receta = row
            ws.append([receta_nombre, total_pedidos_receta, round(total_ingreso_receta, 2)])

            # Acumuladores para el total
            total_pedidos += total_pedidos_receta
            total_ingreso += total_ingreso_receta

        # 🔹 Fila de totales en negrita
        last_row = ws.max_row + 1
        ws.append(["TOTAL", total_pedidos, round(total_ingreso, 2)])

        for cell in ws[last_row]:
            cell.font = Font(bold=True, size=12)

        # 🔹 Ajustar tamaño de columnas automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # 🔹 Guardar el archivo Excel
        file_path = f"reporte_pedidos_{fecha_desde}_a_{fecha_hasta}.xlsx"
        wb.save(file_path)

        conn.close()

        return send_file(file_path, as_attachment=True)

    except sqlite3.Error as e:
        print("❌ Error al generar el reporte:", e)
        return jsonify({"error": "Error al generar el reporte"}), 500

    

if __name__ == '__main__':
    app.run(debug=True)